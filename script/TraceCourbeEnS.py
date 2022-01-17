import numpy
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
from script.utils import DATES, SHEETS

# apparament python peut pas te dire si l'objet est numérique pour un float
def isNumber(s):
    try:
        if s is not None:
            complex(s) # for int, long, float and complex
        else:
            return False
    except ValueError:
        return False

    return True


def findWeek(dates = None, start = DATES["SEMESTER_START"], periods = 16):
    if dates is None:
        dates = pd.date_range(start=start, periods=periods, freq="7D") #trimeste = 16 semaines????

    # Find current week
    week = 0
    for _, date in enumerate(dates):
        if date > DATES["TODAY"]:
            break
        week += 1
    return week


class CourbeEnS():
    previsionSheetName = "Prevision_Courbe_S"
    progressSheetName = "Avancement_Courbe_S"
    workSheetName = "Travail_Courbe_S"

    def __init__(self, excelFile):
        self.wb = openpyxl.load_workbook(excelFile, data_only=True)

        heurePrevue = []
        heureProgress = []
        heureTravailler = []

        if(SHEETS["CBTP"] in self.wb.sheetnames):
            previsionSheet = self.wb[SHEETS["CBTP"]]
            heurePrevue = self.heurePrevue(previsionSheet)

        if(SHEETS["CBTE"] in self.wb.sheetnames):
            progressSheet = self.wb[SHEETS["CBTE"]]
            heureProgress = self.heureProgress(progressSheet)

        if(SHEETS["CRTE"] in self.wb.sheetnames):
            workSheet = self.wb[SHEETS["CRTE"]]
            heureTravailler = self.heureTravailler(workSheet)

        self.genereGraphique(heureProgress, heurePrevue, heureTravailler)

    def heurePrevue(self, feuille):
        colD_valeur = list(feuille.iter_cols(min_col=5, max_col=5, min_row=1, max_row=1000, values_only=True))[0]
        index = colD_valeur.index(max(cellule for cellule in colD_valeur if isNumber(cellule)))+1
        heurePrevueParSemaine = list()
        for cell in feuille[F"F{index}":F"U{index}"][0]:
            heurePrevueParSemaine.append(cell.value)
        print(heurePrevueParSemaine)
        return heurePrevueParSemaine

    def heureProgress(self, feuille):
        colB_valeur = list(feuille.iter_cols(min_col=2, max_col=2, min_row=1, max_row=1000, values_only=True))[0]
        index = colB_valeur.index("TOTAL")+1
        
        heureAvancement = list()
        for cell in feuille[F"F{index}":F"U{index}"][0]:
            heureAvancement.append(cell.value)
        print(heureAvancement)
        return heureAvancement 

    def heureTravailler(self, feuille):
        #iter_cols retourne un itérateur. Convertie l'itérateur en list (pour faire des recherches facilement) pis prendre le premier tuple dedans
        colB_valeur = list(feuille.iter_cols(min_col=2, max_col=2, min_row=1, max_row=50, values_only=True))[0]
        index = colB_valeur.index("Total")+1
        
        heureParSemaine = list()
        for cell in feuille[F"D{index}":F"S{index}"][0]:
            heureParSemaine.append(cell.value)
        print(heureParSemaine)
        return heureParSemaine


    def genereGraphique(self, realProgressHours, BudgetedHours, workedHours):
        #range de date (axe X)
        _, ax = plt.subplots()
        dates = pd.date_range(start=DATES["SEMESTER_START"], periods=len(BudgetedHours), freq="7D")
        week = findWeek(dates=dates)
        ax.plot(BudgetedHours, "b")

        #jusqu'à la semaine 8, worked hours = realprogress = budgeted hour
        #TODO: retirer pour prochaine session
        # hoursOffset = numpy.array((BudgetedHours[:8] + ([BudgetedHours[8]]*(16-8))))

        ax.plot(range(week+1), workedHours[:week+1], "r--")
        ax.plot(range(week+1), realProgressHours[:week+1], "g")

        #deltaAvancement = BudgetedHours[index] - workedHours[index]
        #ax.plot(range(index, len(BudgetedHours)), list((heureTotal-deltaAvancement) for heureTotal in BudgetedHours[index:]), "r--")
        
        ax.set_xticks(range(len(BudgetedHours)))
        ax.set_xticklabels(dates.strftime("%Y-%m-%d"))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right", rotation_mode="anchor")
        ax.legend(("CBTP : Heures totales", "CRTE : Heures travaillées", "CBTE : Heures acquises"))
        #plt.show()
        plt.savefig('img/Courbe_S.pdf', bbox_inches='tight', dpi=96)
